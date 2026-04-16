"""Build docs/project-plan-production.xlsx — sheet 1 = full plan, sheet 2 = RACI."""
from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Install openpyxl: python3 -m pip install openpyxl")

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "project-plan-production.xlsx"

HEADERS = ["Section", "Topic", "Detail", "Phase (indicative)", "Notes / owner"]


def auto_width(ws, max_cap: int = 72):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = len(str(col[0].value or ""))
        for cell in col[1:]:
            if cell.value is None:
                continue
            max_len = max(max_len, min(len(str(cell.value)), max_cap))
        ws.column_dimensions[letter].width = min(max_cap, max(14, max_len + 2))


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Project_plan"

    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")

    plan_rows: list[tuple[str, str, str, str, str]] = [
        (
            "Document",
            "Title",
            "Allege prototype → production roadmap (single sheet). Companion: docs/project-plan-production.md, docs/visual-overview.html.",
            "",
            "",
        ),
        (
            "Goals",
            "Ingestion",
            "Prototype: .eml file drop. Production: continuous Nomura mailbox ingestion (connector + queue + workers).",
            "B",
            "",
        ),
        (
            "Goals",
            "Case identity",
            "Prototype: hash-style demo IDs. Production: stable IDs from enterprise message / thread / conversation scheme.",
            "B",
            "",
        ),
        (
            "Goals",
            "Booking truth",
            "Prototype: BO/MO/FO CSVs. Production: authoritative APIs or certified feeds; interface control documents (ICDs).",
            "C",
            "",
        ),
        (
            "Goals",
            "Outbound",
            "Prototype: sent/ folder. Production: approved SMTP/Graph send with audit (body hash, timestamp, sender identity).",
            "B–D",
            "",
        ),
        (
            "Goals",
            "Access & ops",
            "SSO/RBAC, HA persistence, backups, monitoring, runbooks vs local SQLite / single-node demo.",
            "A, E",
            "",
        ),
        (
            "Principles",
            "Human-in-the-loop",
            "Settlement-impacting actions stay analyst-approved until policy allows automation.",
            "All",
            "",
        ),
        (
            "Principles",
            "Deterministic core",
            "Pipeline classify → extract → match → draft remains testable; LLM augments with schema validation, not sole truth for booking keys.",
            "All",
            "",
        ),
        (
            "Principles",
            "Integration by contracts",
            "Versioned APIs/events from mail and booking platforms; avoid long-term screen-scrape bridges.",
            "B–C",
            "",
        ),
        (
            "Principles",
            "Security by design",
            "Secrets in vault, least-privilege service accounts, encryption in transit and at rest, redaction for logs/LLM egress.",
            "A",
            "Security",
        ),
        (
            "Phases",
            "A — Foundation",
            "Vault/secrets, CI/CD, env promotion, structured logs + metrics + tracing, SSO + RBAC baseline.",
            "4–8 wk",
            "Eng + Sec",
        ),
        (
            "Phases",
            "B — Mailbox MVP",
            "Connector (poll/subscribe), idempotency, queue to pipeline workers, non-prod Nomura mailboxes, outbound UAT path.",
            "8–14 wk",
            "Eng + Platform",
        ),
        (
            "Phases",
            "C — Booking & SSI",
            "Replace CSV with read APIs/feeds for BO→MO→FO (or agreed order); SSI/static data source; contract tests on each release.",
            "10–16 wk",
            "Overlap B",
        ),
        (
            "Phases",
            "D — Pilot",
            "Limited desk, KPIs (time-to-touch, precision/recall sample, incidents), parallel or shadow vs legacy.",
            "4–6 wk",
            "Ops + Eng",
        ),
        (
            "Phases",
            "E — GA + hardening",
            "Scale workers, DR drills, SLOs, on-call runbooks; optional agent layer only after D stable.",
            "ongoing",
            "",
        ),
        (
            "Nomura mailbox",
            "Discovery",
            "Which shared mailboxes in v1; Microsoft 365 Graph vs on-prem EWS; retention/journaling/DLP; no production-mail model training (typical policy).",
            "A–B",
            "Ops + Sec",
        ),
        (
            "Nomura mailbox",
            "Ingestion",
            "Normalize to internal message DTO; dedup/replay-safe message id; dead-letter + retry for connector failures.",
            "B",
            "",
        ),
        (
            "Nomura mailbox",
            "Threading",
            "Align Message-ID, In-Reply-To, References, conversation IDs to internal case; define 'latest reply' vs full thread for classifier.",
            "B",
            "",
        ),
        (
            "Nomura mailbox",
            "Outbound",
            "Nominated send identity; pre-send review in UI; correlate send to case and original inbound message.",
            "B–D",
            "",
        ),
        (
            "Nomura mailbox",
            "Non-functional",
            "API rate limits, attachment size limits, malware scan before extract, PII handling for third-party LLM.",
            "B",
            "Sec",
        ),
        (
            "Internal systems",
            "Murex (or primary book)",
            "Canonical trade economics and refs; REST/SOAP/export per Nomura-approved pattern; matcher keys must match ICD.",
            "C",
            "Platform",
        ),
        (
            "Internal systems",
            "BO / MO / FO",
            "Read-only integration first; preserve business rule BO→MO→FO stop-at-first-hit if still mandated.",
            "C",
            "",
        ),
        (
            "Internal systems",
            "SSI / static data",
            "BIC, agents, standing instructions for mismatch reasons and drafter context.",
            "C",
            "",
        ),
        (
            "Internal systems",
            "Channels (DTCC / MarkitWire / SWIFT)",
            "Clarify whether pipeline reads only from Murex/hub vs multiple hops; reduce duplicate truth.",
            "C",
            "Architecture",
        ),
        (
            "Internal systems",
            "IAM",
            "Entra ID / LDAP SSO; JWT on APIs; desk/entity entitlements from directory groups.",
            "A",
            "IAM team",
        ),
        (
            "Internal systems",
            "Audit",
            "Append-only store; link user id, case id, message id for regulatory samples.",
            "A–E",
            "",
        ),
        (
            "Internal systems",
            "Ticketing (optional)",
            "ServiceNow/Jira webhook from Escalate actions for FO/SSI workflows.",
            "D+",
            "",
        ),
        (
            "LLM governance",
            "Hosting",
            "Cloud API vs on-prem / Hugging Face endpoints — network zone, latency SLO, data residency.",
            "A–C",
            "Sec + Eng",
        ),
        (
            "LLM governance",
            "Controls",
            "Prompt versioning, PII scrubbing, JSON schema validation on outputs, low-confidence surfacing in UI.",
            "A–D",
            "",
        ),
        (
            "LLM governance",
            "Fallback",
            "Rules-only classifier/extract path when LLM disabled or unhealthy (align with prototype rules path).",
            "All",
            "",
        ),
        (
            "Risks",
            "Mail API throttling",
            "Mitigation: back-pressure, batching, off-peak catch-up jobs.",
            "B",
            "",
        ),
        (
            "Risks",
            "Field drift",
            "Mitigation: versioned contracts with Murex/feeds; automated contract tests on release.",
            "C",
            "",
        ),
        (
            "Risks",
            "LLM extract errors",
            "Mitigation: schema validation, human review band, confidence thresholds.",
            "B–D",
            "",
        ),
        (
            "Risks",
            "Scope creep",
            "Mitigation: vertical slice — one product, one desk, one mailbox before broad rollout.",
            "All",
            "SteerCo",
        ),
        (
            "Exit checklist",
            "ICDs published",
            "Mailbox, primary booking, SSI — signed interface control documents.",
            "Pre-pilot",
            "Y/N",
        ),
        (
            "Exit checklist",
            "Non-prod E2E",
            "Ingest → case → analyst review → send on UAT relay (full chain).",
            "Pre-pilot",
            "Y/N",
        ),
        (
            "Exit checklist",
            "SSO + RBAC",
            "Production identity integrated; roles mapped to groups.",
            "Pre-pilot",
            "Y/N",
        ),
        (
            "Exit checklist",
            "Audit export",
            "Sample export passes compliance review.",
            "Pre-pilot",
            "Y/N",
        ),
        (
            "Exit checklist",
            "DR",
            "RPO/RTO agreed and tested for case DB + queue.",
            "Pre-GA",
            "Y/N",
        ),
        (
            "Exit checklist",
            "Runbooks",
            "Connector down, booking down, LLM down — on-call procedures.",
            "GA",
            "Y/N",
        ),
    ]

    for row in plan_rows:
        ws.append(list(row))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    auto_width(ws)

    # --- Sheet 2: RACI ---
    raci = wb.create_sheet("RACI")
    raci.append(["Workstream", "Engineering", "Business / Ops", "Security", "Platform (mail / Murex)"])
    for c in raci[1]:
        c.font = Font(bold=True)
    for r in [
        ("Mailbox connector", "R / A", "C", "C", "A"),
        ("Pipeline / matcher", "R / A", "C", "I", "C"),
        ("UI / workflows", "R / A", "C", "C", "I"),
        ("LLM governance", "R", "I", "A", "I"),
        ("SSO / IAM", "C", "I", "A", "R"),
    ]:
        raci.append(list(r))
    raci.append([])
    raci.append(["Legend:", "R = Responsible", "A = Accountable", "C = Consulted", "I = Informed"])
    for row in raci.iter_rows(min_row=2, max_row=raci.max_row - 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    auto_width(raci, max_cap=40)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
